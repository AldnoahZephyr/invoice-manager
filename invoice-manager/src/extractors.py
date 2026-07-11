# -*- coding: utf-8 -*-
"""
提取器模块
负责从 PDF 电子发票和 Excel 文件中提取发票号码与金额
"""

import os
import re
import sys


# ======================================================================
# PDF 电子发票提取
# ======================================================================

def extract_invoice_number(text):
    """
    从文本中提取发票号码
    支持全电发票（20位）和传统电子发票（8位）

    PDF文本提取经常出现碎片化：标签"发票号码："和实际号码之间可能
    隔了很多无关行，因此不能简单地用 [^\\d]* 跨行匹配，否则会误抓
    "下载次数：11"等干扰数字。
    """
    text = text or ''

    # ---- 策略1: "发票号码" 标签后紧跟数字（不跨行） ----
    for pattern in [
        r'发票号码[：:\s]*[^\d\n]{0,5}(\d{20})',   # 全电发票 20 位
        r'发票号码[：:\s]*[^\d\n]{0,5}(\d{8})\b',   # 传统电子发票 8 位
    ]:
        match = re.search(pattern, text)
        if match:
            return match.group(1)

    # ---- 策略2: 查找20位纯数字（全电发票号码，在文本中通常唯一） ----
    match = re.search(r'(?<!\d)(\d{20})(?!\d)', text)
    if match:
        return match.group(1)

    # ---- 策略3: "制" / "监制" 后面的数字（发票监制章区域常见） ----
    match = re.search(r'制\s*[：:\s]*(\d{8,20})', text)
    if match:
        return match.group(1)

    # ---- 策略4: "发票号码" 后有限范围内查找8位数字（限制距离避免跨行干扰） ----
    match = re.search(r'发票号码[：:\s]*[^\d]{0,20}?(\d{8})\b', text)
    if match:
        num = match.group(1)
        # 排除 "下载次数" / "页码" / "序号" 等干扰词后的数字
        prefix = text[max(0, match.start() - 10):match.start()]
        if not re.search(r'下载|次数|页码|序号|版次', prefix):
            return num

    # ---- 策略5: 兜底 - 查找8位以上纯数字，排除日期和信用代码片段 ----
    for m in re.finditer(r'(?<!\d)(\d{8,19})(?!\d)', text):
        num = m.group(1)
        # 排除日期格式（如 20260702）
        if re.match(r'^20\d{6}$', num):
            continue
        # 排除信用代码片段（紧邻字母的是统一社会信用代码的一部分）
        ctx_before = text[max(0, m.start() - 3):m.start()]
        ctx_after = text[m.end():m.end() + 3]
        if re.search(r'[A-Za-z]$', ctx_before) or re.search(r'^[A-Za-z]', ctx_after):
            continue
        return num

    return ''


def extract_amount(text):
    """
    从文本中提取金额
    优先匹配 "合计" / "价税合计"，其次 "金额"，最后匹配 ¥/￥/CNY 符号后的数值
    """
    text = text or ''

    # 辅助函数：清理并转换金额字符串
    def _to_float(s):
        s = s.replace(',', '').replace('，', '').replace(' ', '')
        try:
            return float(s)
        except ValueError:
            return None

    patterns = [
        # 合计金额（机票行程单等常见）
        # 允许 "合计" 与金额之间跨行，支持 CNY/¥/￥ 前缀
        r'合计[（(]金额[）)]?[：:\s]*(?:CNY|¥|￥)?\s*([\d,]+\.\d{2})',
        r'合计[（(]金额[）)]?[：:\s]*(?:CNY|¥|￥)?\s*([\d,]+)',
        r'合计[：:\s]*(?:CNY|¥|￥)?\s*([\d,]+\.\d{2})',
        r'合计[：:\s]*(?:CNY|¥|￥)?\s*([\d,]+)',
        # 总计 / 总金额
        r'(?:总[金全]额|总计)[：:\s]*(?:CNY|¥|￥)?\s*([\d,]+\.\d{2})',
        r'(?:总[金全]额|总计)[：:\s]*(?:CNY|¥|￥)?\s*([\d,]+)',
        # 金额：¥123.45 / 金额: 123.45 / 金额 ￥1,234.56
        r'金额[：:\s]*(?:CNY|¥|￥)?\s*([\d,]+\.\d{2})',
        r'金额[：:\s]*(?:CNY|¥|￥)?\s*([\d,]+)',
        # 价税合计（大写金额后的数字）
        r'价税合计[（(]大写[）)][^¥￥CNY]*(?:CNY|¥|￥)?\s*([\d,]+\.\d{2})',
        r'价税合计[：:\s]*(?:CNY|¥|￥)?\s*([\d,]+\.\d{2})',
        # ¥/￥/CNY 符号后的金额
        r'(?:CNY|¥|￥)\s*([\d,]+\.\d{2})',
    ]
    all_valid_amounts = []
    for pattern in patterns:
        # 收集当前规则下的所有匹配值，避免同规则下的大额值被漏掉
        matches = re.findall(pattern, text)
        for match_str in matches:
            amount = _to_float(match_str)
            if amount is not None and amount > 0:
                all_valid_amounts.append(amount)

    # 所有匹配结果里取最大值，无有效结果则返回兜底0.0
    return max(all_valid_amounts) if all_valid_amounts else 0.0


# ======================================================================
# OCR 兜底：处理图片型 PDF（如机票行程单）
# ======================================================================

# 模块级自定义 Tesseract 路径（由 set_tesseract_cmd 设置，优先级最高）
_CUSTOM_TESSERACT_CMD = None


def set_tesseract_cmd(path):
    """
    设置自定义 Tesseract 可执行文件路径。
    设置后 _find_tesseract_cmd() 会优先使用此路径。
    传入空字符串或 None 则清除自定义路径，恢复自动查找。
    """
    global _CUSTOM_TESSERACT_CMD
    if path and path.strip():
        _CUSTOM_TESSERACT_CMD = path.strip()
    else:
        _CUSTOM_TESSERACT_CMD = None


def get_tesseract_cmd():
    """获取当前生效的 Tesseract 路径（可能为 None）"""
    return _find_tesseract_cmd()


def _find_tesseract_cmd():
    """
    查找 Tesseract 可执行文件路径。
    查找顺序：
      0. 模块级自定义路径（由 set_tesseract_cmd 设置，优先级最高）
      1. 环境变量 TESSERACT_CMD
      2. 项目目录下的 tesseract/tesseract.exe
      3. 系统 PATH 中的 tesseract
      4. 常见 Windows 安装路径
    """
    import shutil

    # 优先检查自定义路径
    if _CUSTOM_TESSERACT_CMD and os.path.isfile(_CUSTOM_TESSERACT_CMD):
        return _CUSTOM_TESSERACT_CMD

    candidates = []

    env_cmd = os.environ.get('TESSERACT_CMD')
    if env_cmd and os.path.isfile(env_cmd):
        candidates.append(env_cmd)

    # 打包后 tesseract 在 sys._MEIPASS/tesseract/；开发时在项目根目录/tesseract/
    if getattr(sys, 'frozen', False):
        project_root = sys._MEIPASS
    else:
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    bundled = os.path.join(project_root, 'tesseract', 'tesseract.exe')
    if os.path.isfile(bundled):
        candidates.append(bundled)

    system_tess = shutil.which('tesseract')
    if system_tess:
        candidates.append(system_tess)

    common_paths = [
        r'C:\Program Files\Tesseract-OCR\tesseract.exe',
        r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
    ]
    for p in common_paths:
        if os.path.isfile(p):
            candidates.append(p)

    # 去重（保持顺序）
    seen = set()
    for cmd in candidates:
        if cmd and cmd not in seen:
            seen.add(cmd)
            return cmd
    return None


def _ocr_available():
    """检查 OCR 依赖是否可用"""
    try:
        import pytesseract  # noqa: F401
        import fitz          # noqa: F401
        from PIL import Image  # noqa: F401
        return _find_tesseract_cmd() is not None
    except Exception:
        return False


def _render_page_to_image(page, zoom=4):
    """将 PDF 页面渲染为 PIL Image"""
    from PIL import Image
    mat = page.transformation_matrix if hasattr(page, 'transformation_matrix') else None
    # 使用 Matrix 放大，提高 OCR 精度
    matrix = page.get_matrix() if hasattr(page, 'get_matrix') else None
    if matrix is None:
        # PyMuPDF 较新版本用法
        import fitz
        matrix = fitz.Matrix(zoom, zoom)
    else:
        import fitz
        matrix = fitz.Matrix(zoom, zoom)
    pix = page.get_pixmap(matrix=matrix)
    img = Image.frombytes('RGB', [pix.width, pix.height], pix.samples)
    return img


def _enhance_for_ocr(image):
    """对图像做简单预处理以提升 OCR 效果"""
    from PIL import ImageEnhance
    gray = image.convert('L')
    enhancer = ImageEnhance.Contrast(gray)
    return enhancer.enhance(2.0)


def _ocr_image(image, tesseract_cmd, lang='chi_sim+eng'):
    """对单张图片执行 OCR，返回识别到的文本"""
    import pytesseract
    pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
    try:
        return pytesseract.image_to_string(image, lang=lang) or ''
    except Exception as e:
        print(f'[OCR 错误]: {e}')
        return ''


def _extract_text_with_ocr(pdf_path):
    """
    对图片型 PDF 使用 OCR 提取文本。
    为提高关键字段命中率，会同时识别：
      - 整页（overview）
      - 右上角区域（发票号码常见位置）
      - 右下角区域（合计金额常见位置）
    返回合并后的文本。
    """
    import fitz

    tesseract_cmd = _find_tesseract_cmd()
    if not tesseract_cmd:
        return ''

    text_parts = []
    try:
        doc = fitz.open(pdf_path)
        for page in doc:
            width = page.rect.width
            height = page.rect.height

            # 1) 整页 3x，用于获取整体内容
            full_img = _render_page_to_image(page, zoom=3)
            full_img = _enhance_for_ocr(full_img)
            text_parts.append(_ocr_image(full_img, tesseract_cmd))

            # 2) 右上角 4x（发票号码、开票状态等）
            #    取顶部 30%、右侧 45% 区域
            tr_crop = full_img.crop((int(width * 3 * 0.55), 0, int(width * 3), int(height * 3 * 0.30)))
            tr_crop = _enhance_for_ocr(tr_crop)
            text_parts.append(_ocr_image(tr_crop, tesseract_cmd))

            # 3) 右下角 4x（合计金额、税费等）
            #    取底部 45%、右侧 55% 区域
            br_crop = full_img.crop((int(width * 3 * 0.45), int(height * 3 * 0.55), int(width * 3), int(height * 3)))
            br_crop = _enhance_for_ocr(br_crop)
            text_parts.append(_ocr_image(br_crop, tesseract_cmd))

        doc.close()
    except Exception as e:
        print(f'[OCR PDF 错误] {pdf_path}: {e}')

    return '\n'.join(text_parts)


def extract_pdf_invoice(pdf_path):
    """
    从 PDF 文件中提取发票号码和金额
    返回: (invoice_number: str, amount: float)
    """
    try:
        import pdfplumber
        text = ''
        needs_ocr = False

        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + '\n'
                # 如果页面没有任何文本字符但包含图片，则很可能是图片型 PDF
                if len(page.chars) == 0 and len(page.images) > 0:
                    needs_ocr = True

        # 清理文本中的多余空白，但保留关键结构
        clean_text = re.sub(r'[ \t]+', ' ', text)
        invoice_number = extract_invoice_number(clean_text)
        amount = extract_amount(clean_text)

        # 如果 pdfplumber 没识别到关键信息，且页面是图片型，则启用 OCR 兜底
        if (not invoice_number or not amount or needs_ocr) and _ocr_available():
            ocr_text = _extract_text_with_ocr(pdf_path)
            ocr_text = re.sub(r'[ \t]+', ' ', ocr_text)
            # OCR 结果与原始文本合并后再次提取
            combined_text = clean_text + '\n' + ocr_text
            ocr_invoice = extract_invoice_number(combined_text)
            ocr_amount = extract_amount(combined_text)
            # 优先使用识别到的非空结果
            invoice_number = invoice_number or ocr_invoice
            amount = amount or ocr_amount

        return invoice_number, amount
    except Exception as e:
        print(f'[PDF提取错误] {pdf_path}: {e}')
        return '', 0.0


def scan_pdf_folder(folder_path):
    """
    扫描文件夹下所有 PDF 文件，提取发票信息
    返回: list of dict，每个 dict 包含 file_name, invoice_number, amount, file_path
    """
    results = []
    if not os.path.isdir(folder_path):
        return results
    for filename in sorted(os.listdir(folder_path)):
        if filename.lower().endswith('.pdf'):
            pdf_path = os.path.join(folder_path, filename)
            invoice_number, amount = extract_pdf_invoice(pdf_path)
            results.append({
                'file_name': filename,
                'invoice_number': invoice_number,
                'amount': amount,
                'file_path': pdf_path,
            })
    return results


# ======================================================================
# OFD 电子发票提取
# ======================================================================
# OFD（Open Fixed-layout Document）是中国版式文档国家标准。
# OFD 文件本质是一个 ZIP 压缩包，内部由 XML 文件和图片资源组成。
# 典型结构：
#   OFD.xml            → 根文件，指向 Document.xml
#   Doc_0/Document.xml → 文档描述，列出所有页面
#   Doc_0/Pages/Page_0/Content.xml → 页面内容，包含 <TextCode> 文本
#   Doc_0/Res/         → 图片等资源
# ======================================================================

# OFD 标准命名空间
_OFD_NS = 'http://www.ofdspec.org/2016'


def _ofd_xml_parse(zip_file, xml_path):
    """从 ZIP 中读取并解析 XML 文件"""
    try:
        raw = zip_file.read(xml_path)
        import xml.etree.ElementTree as ET
        return ET.fromstring(raw)
    except Exception:
        return None


def _extract_ofd_text_from_xml(element):
    """
    递归提取 XML 元素中所有文本内容。
    优先提取 <TextCode> 元素（OFD 中的文本对象），
    同时也收集其他元素的文本作为兜底。
    """
    texts = []

    if element is None:
        return texts

    tag = element.tag.split('}')[-1] if '}' in element.tag else element.tag

    # TextCode 是 OFD 中存放实际文字的元素
    if tag == 'TextCode':
        if element.text:
            texts.append(element.text)
        if element.tail:
            texts.append(element.tail)

    # 递归处理子元素
    for child in element:
        texts.extend(_extract_ofd_text_from_xml(child))

    # 非 TextCode 元素的直接文本（兜底）
    if tag != 'TextCode' and element.text and element.text.strip():
        texts.append(element.text)

    return texts


def _resolve_ofd_path(base_dir, ref):
    """
    解析 OFD 内部引用路径（相对于 base_dir）。
    OFD 中 BaseLoc 属性可能是相对路径，需要拼接。
    """
    ref = ref.strip()
    # 去掉开头的 ./
    if ref.startswith('./'):
        ref = ref[2:]

    # 尝试直接拼接
    candidate = os.path.normpath(os.path.join(base_dir, ref)).replace('\\', '/')
    return candidate


def _extract_ofd_images_for_ocr(zip_file, doc_dir):
    """
    从 OFD 中提取图片资源用于 OCR。
    扫描 ZIP 内所有图片文件（jpg/png/bmp/tif），返回 PIL Image 列表。
    """
    from PIL import Image
    import io

    image_exts = ('.jpg', '.jpeg', '.png', '.bmp', '.tif', '.tiff')
    images = []
    for name in zip_file.namelist():
        if name.lower().endswith(image_exts):
            try:
                data = zip_file.read(name)
                img = Image.open(io.BytesIO(data))
                images.append(img.convert('RGB'))
            except Exception:
                pass
    return images


def extract_ofd_invoice(ofd_path):
    """
    从 OFD 文件中提取发票号码和金额
    OFD 是 ZIP+XML 格式，通过解析内部 XML 提取文本。
    若文本提取失败（图片型 OFD），则提取图片做 OCR 兜底。
    返回: (invoice_number: str, amount: float)
    """
    try:
        import zipfile

        all_text = []
        needs_ocr = False

        with zipfile.ZipFile(ofd_path, 'r') as zf:
            names = zf.namelist()

            # ---- 1. 解析 OFD.xml 找到文档根路径 ----
            doc_roots = []
            ofd_xml = _ofd_xml_parse(zf, 'OFD.xml')
            if ofd_xml is not None:
                # 查找 DocRoot 元素的 DocRoot 属性
                for elem in ofd_xml.iter():
                    tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
                    if tag == 'DocRoot':
                        loc = elem.get('DocRoot', '')
                        if loc:
                            doc_roots.append(loc)

            # ---- 2. 解析每个文档的 Document.xml，找到页面路径 ----
            page_paths = []
            for doc_root in doc_roots:
                doc_dir = os.path.dirname(doc_root)
                doc_xml = _ofd_xml_parse(zf, doc_root)
                if doc_xml is not None:
                    for elem in doc_xml.iter():
                        tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
                        if tag == 'Page':
                            base_loc = elem.get('BaseLoc', '')
                            if base_loc:
                                page_path = _resolve_ofd_path(doc_dir, base_loc)
                                page_paths.append(page_path)

            # ---- 3. 解析每个页面的 Content.xml，提取文本 ----
            if page_paths:
                for page_path in page_paths:
                    page_xml = _ofd_xml_parse(zf, page_path)
                    if page_xml is not None:
                        texts = _extract_ofd_text_from_xml(page_xml)
                        all_text.extend(texts)
                    else:
                        needs_ocr = True
            else:
                # 如果没找到页面路径，尝试暴力解析所有 XML
                needs_ocr = True

            # 暴力兜底：解析所有 XML 文件中的文本
            for name in names:
                if name.lower().endswith('.xml'):
                    xml_elem = _ofd_xml_parse(zf, name)
                    if xml_elem is not None:
                        texts = _extract_ofd_text_from_xml(xml_elem)
                        all_text.extend(texts)

            # ---- 4. 合并文本并提取发票信息 ----
            combined_text = '\n'.join(all_text)
            combined_text = re.sub(r'[ \t]+', ' ', combined_text)
            invoice_number = extract_invoice_number(combined_text)
            amount = extract_amount(combined_text)

            # ---- 5. 如果没识别到关键信息，启用 OCR 兜底 ----
            if (not invoice_number or not amount) and _ocr_available():
                doc_dir = os.path.dirname(doc_roots[0]) if doc_roots else ''
                images = _extract_ofd_images_for_ocr(zf, doc_dir)
                if images:
                    tesseract_cmd = _find_tesseract_cmd()
                    ocr_parts = []
                    for img in images:
                        enhanced = _enhance_for_ocr(img)
                        ocr_parts.append(_ocr_image(enhanced, tesseract_cmd))
                    ocr_text = '\n'.join(ocr_parts)
                    ocr_text = re.sub(r'[ \t]+', ' ', ocr_text)
                    full_text = combined_text + '\n' + ocr_text
                    ocr_invoice = extract_invoice_number(full_text)
                    ocr_amount = extract_amount(full_text)
                    invoice_number = invoice_number or ocr_invoice
                    amount = amount or ocr_amount

        return invoice_number, amount
    except Exception as e:
        print(f'[OFD提取错误] {ofd_path}: {e}')
        return '', 0.0


# ======================================================================
# 统一发票提取入口（自动识别 PDF / OFD）
# ======================================================================

def extract_invoice(file_path):
    """
    从发票文件中提取发票号码和金额（自动识别 PDF 或 OFD 格式）
    返回: (invoice_number: str, amount: float)
    """
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.pdf':
        return extract_pdf_invoice(file_path)
    elif ext == '.ofd':
        return extract_ofd_invoice(file_path)
    return '', 0.0


def scan_invoice_folder(folder_path):
    """
    扫描文件夹下所有 PDF 和 OFD 文件，提取发票信息
    返回: list of dict，每个 dict 包含 file_name, invoice_number, amount, file_path
    """
    results = []
    if not os.path.isdir(folder_path):
        return results
    for filename in sorted(os.listdir(folder_path)):
        ext = filename.lower()
        if ext.endswith('.pdf') or ext.endswith('.ofd'):
            file_path = os.path.join(folder_path, filename)
            invoice_number, amount = extract_invoice(file_path)
            results.append({
                'file_name': filename,
                'invoice_number': invoice_number,
                'amount': amount,
                'file_path': file_path,
            })
    return results


# scan_pdf_folder 作为兼容别名保留
# scan_pdf_folder 已在上方定义，scan_invoice_folder 为新版统一扫描函数


# ======================================================================
# Excel 文件提取
# ======================================================================

def _read_excel_rows(excel_path):
    """读取 Excel 文件，返回行列表（每行为值列表）"""
    ext = os.path.splitext(excel_path)[1].lower()
    if ext == '.xlsx':
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()
        return rows
    elif ext == '.xls':
        import xlrd
        wb = xlrd.open_workbook(excel_path)
        ws = wb.sheet_by_index(0)
        rows = [ws.row_values(i) for i in range(ws.nrows)]
        return rows
    else:
        raise ValueError(f'不支持的文件格式: {ext}')


def extract_excel_invoice(excel_path):
    """
    从 Excel 文件中提取发票号码和金额
    要求 Excel 中存在列名为 "发票号码" 和 "金额" 的列
    返回: list of dict，每个 dict 包含 file_name, invoice_number, amount
    """
    results = []
    try:
        rows = _read_excel_rows(excel_path)
        if not rows:
            return results

        # 在前几行中查找列名
        header_row_idx = None
        invoice_col = None
        amount_col = None
        for row_idx, row in enumerate(rows[:10]):  # 最多检查前 10 行
            for col_idx, cell in enumerate(row):
                cell_str = str(cell).strip() if cell is not None else ''
                if '发票号码' in cell_str and invoice_col is None:
                    invoice_col = col_idx
                    header_row_idx = row_idx
                if '金额' in cell_str and amount_col is None:
                    amount_col = col_idx
                    header_row_idx = row_idx
            if invoice_col is not None and amount_col is not None:
                break

        if invoice_col is None or amount_col is None:
            return results

        # 从表头行的下一行开始读取数据
        for row in rows[header_row_idx + 1:]:
            invoice_number = ''
            if invoice_col < len(row) and row[invoice_col] is not None:
                invoice_number = str(row[invoice_col]).strip()
                # 去除可能的小数点（Excel 数字格式化问题）
                if invoice_number.endswith('.0'):
                    invoice_number = invoice_number[:-2]

            amount = 0.0
            if amount_col < len(row) and row[amount_col] is not None:
                try:
                    amount = float(row[amount_col])
                except (ValueError, TypeError):
                    pass

            if invoice_number or amount:
                results.append({
                    'file_name': '',
                    'invoice_number': invoice_number,
                    'amount': amount,
                    'file_path': '',
                })
    except Exception as e:
        print(f'[Excel提取错误] {excel_path}: {e}')
    return results
