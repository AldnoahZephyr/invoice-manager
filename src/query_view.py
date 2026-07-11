# -*- coding: utf-8 -*-
"""
入库查询视图模块
实现查询、导出、批量编辑、批量删除等功能
"""

from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QLineEdit, QFileDialog, QMessageBox, QDialog, QFormLayout,
    QGridLayout, QHeaderView
)

from .widgets import CheckBoxTable, TableToolbar
from .database import Database


class BatchEditDialog(QDialog):
    """批量编辑对话框"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._init_ui()

    def _init_ui(self):
        self.setWindowTitle('批量编辑')
        self.setMinimumWidth(400)

        layout = QVBoxLayout(self)

        lbl = QLabel('将选中的记录统一修改以下字段（留空表示不修改）：')
        lbl.setStyleSheet('padding: 8px;')
        layout.addWidget(lbl)

        form = QFormLayout()
        self.edt_type = QLineEdit()
        self.edt_process = QLineEdit()
        self.edt_time = QLineEdit()
        self.edt_dept = QLineEdit()
        self.edt_name = QLineEdit()
        self.edt_remark = QLineEdit()

        form.addRow('发票种类：', self.edt_type)
        form.addRow('流程编号：', self.edt_process)
        form.addRow('时间：', self.edt_time)
        form.addRow('部门：', self.edt_dept)
        form.addRow('姓名：', self.edt_name)
        form.addRow('备注：', self.edt_remark)
        layout.addLayout(form)

        btn_line = QHBoxLayout()
        btn_line.addStretch()
        btn_ok = QPushButton('确定')
        btn_cancel = QPushButton('取消')
        btn_ok.setFixedWidth(100)
        btn_cancel.setFixedWidth(100)
        btn_ok.clicked.connect(self.accept)
        btn_cancel.clicked.connect(self.reject)
        btn_line.addWidget(btn_ok)
        btn_line.addWidget(btn_cancel)
        layout.addLayout(btn_line)

    def get_fields(self):
        """获取非空字段"""
        fields = {}
        val = self.edt_type.text().strip()
        if val:
            fields['发票种类'] = val
        val = self.edt_process.text().strip()
        if val:
            fields['流程编号'] = val
        val = self.edt_time.text().strip()
        if val:
            fields['时间'] = val
        val = self.edt_dept.text().strip()
        if val:
            fields['部门'] = val
        val = self.edt_name.text().strip()
        if val:
            fields['姓名'] = val
        val = self.edt_remark.text().strip()
        if val:
            fields['备注'] = val
        return fields


class QueryView(QWidget):
    """入库查询视图"""

    # 表格列：发票号码、金额、发票种类、文件名、文件路径、部门、姓名、备注
    # 额外有一隐藏的 id 列用于数据库操作
    TABLE_HEADERS = ['发票号码', '金额', '发票种类', '文件名', '文件路径','流程编号','时间', '部门', '姓名', '备注']

    def __init__(self, db: Database, parent=None):
        super().__init__(parent)
        self.db = db
        self._records = []  # 当前查询结果（包含 id）
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)

        # ---- 顶部查询条件区 ----
        cond_grid = QGridLayout()
        cond_grid.setContentsMargins(0, 0, 0, 0)

        self.inputs = {}
        fields = [
            '发票号码', '金额', '发票种类', '文件名', '归档路径',
            '流程编号', '时间', '部门', '姓名', '备注'
        ]
        for i, field in enumerate(fields):
            row = i // 5
            col = (i % 5) * 2
            cond_grid.addWidget(QLabel(field + '：'), row, col)
            edt = QLineEdit()
            edt.setFixedWidth(160)
            cond_grid.addWidget(edt, row, col + 1)
            self.inputs[field] = edt

        layout.addLayout(cond_grid)

        # ---- 按钮行 ----
        btn_line = QHBoxLayout()
        self.btn_query = QPushButton('查询')
        self.btn_export = QPushButton('导出')
        self.btn_batch_edit = QPushButton('批量编辑')
        self.btn_batch_delete = QPushButton('批量删除')

        for btn in [self.btn_query, self.btn_export, self.btn_batch_edit, self.btn_batch_delete]:
            btn.setFixedWidth(100)

        self.btn_query.clicked.connect(self._on_query)
        self.btn_export.clicked.connect(self._on_export)
        self.btn_batch_edit.clicked.connect(self._on_batch_edit)
        self.btn_batch_delete.clicked.connect(self._on_batch_delete)

        btn_line.addWidget(self.btn_query)
        btn_line.addWidget(self.btn_export)
        btn_line.addWidget(self.btn_batch_edit)
        btn_line.addWidget(self.btn_batch_delete)
        btn_line.addStretch()
        layout.addLayout(btn_line)

        # ---- 中间表格 ----
        self.table = CheckBoxTable(self.TABLE_HEADERS, amount_col_idx=1)
        layout.addWidget(self.table, 1)

        # ---- 底部工具栏 ----
        self.toolbar = TableToolbar(self.table)
        layout.addWidget(self.toolbar)

        # 初始加载全部数据
        self._on_query()

    # ------------------------------------------------------------------
    # 事件处理
    # ------------------------------------------------------------------

    def _on_query(self):
        """查询"""
        conditions = {}
        for field, edt in self.inputs.items():
            conditions[field] = edt.text().strip()

        self._records = self.db.query_records(conditions)
        self._refresh_table()

    def _refresh_table(self):
        """刷新表格"""
        rows = []
        for rec in self._records:
            rows.append([
                rec.get('发票号码', ''),
                f"{rec.get('金额', 0):.2f}",
                rec.get('发票种类', ''),
                rec.get('文件名', ''),
                rec.get('归档路径', ''),
                rec.get('流程编号', ''),
                rec.get('时间', ''),
                rec.get('部门', ''),
                rec.get('姓名', ''),
                rec.get('备注', ''),
            ])
        self.table.set_rows(rows)

    def _on_export(self):
        """导出 Excel"""
        if self.table.get_row_count() == 0:
            QMessageBox.information(self, '提示', '表格中没有数据可导出')
            return

        path, _ = QFileDialog.getSaveFileName(
            self, '导出 Excel', 'InvoiceRecords.xlsx',
            'Excel 文件 (*.xlsx)'
        )
        if not path:
            return

        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = '发票记录'

            # 写表头
            headers = self.TABLE_HEADERS
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)

            # 写数据
            for row_idx, rec in enumerate(self._records, 2):
                ws.cell(row=row_idx, column=1, value=rec.get('发票号码', ''))
                ws.cell(row=row_idx, column=2, value=rec.get('金额', 0))
                ws.cell(row=row_idx, column=3, value=rec.get('发票种类', ''))
                ws.cell(row=row_idx, column=4, value=rec.get('文件名', ''))
                ws.cell(row=row_idx, column=5, value=rec.get('归档路径', ''))
                ws.cell(row=row_idx, column=6, value=rec.get('流程编号', ''))
                ws.cell(row=row_idx, column=7, value=rec.get('时间', ''))
                ws.cell(row=row_idx, column=8, value=rec.get('部门', ''))
                ws.cell(row=row_idx, column=9, value=rec.get('姓名', ''))
                ws.cell(row=row_idx, column=10, value=rec.get('备注', ''))

            wb.save(path)
            QMessageBox.information(self, '成功', f'已导出到：\n{path}')
        except Exception as e:
            QMessageBox.critical(self, '错误', f'导出失败：{e}')

    def _on_batch_edit(self):
        """批量编辑"""
        checked = self.table.get_checked_row_indices()
        if not checked:
            QMessageBox.information(self, '提示', '请先勾选需要编辑的记录')
            return

        dialog = BatchEditDialog(self)
        if dialog.exec_() != QDialog.Accepted:
            return

        fields = dialog.get_fields()
        if not fields:
            QMessageBox.information(self, '提示', '未输入任何修改内容')
            return

        # 获取选中记录的数据库 ID
        ids = [self._records[i]['id'] for i in checked if i < len(self._records)]
        self.db.batch_update(ids, fields)
        self._on_query()
        QMessageBox.information(self, '成功', f'已更新 {len(ids)} 条记录')

    def _on_batch_delete(self):
        """批量删除"""
        checked = self.table.get_checked_row_indices()
        if not checked:
            QMessageBox.information(self, '提示', '请先勾选需要删除的记录')
            return

        reply = QMessageBox.question(
            self, '确认删除',
            f'确定要删除选中的 {len(checked)} 条记录吗？',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        ids = [self._records[i]['id'] for i in checked if i < len(self._records)]
        self.db.batch_delete(ids)
        self._on_query()
        QMessageBox.information(self, '成功', f'已删除 {len(ids)} 条记录')
